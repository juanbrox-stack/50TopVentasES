import streamlit as st
import pandas as pd
import io
from datetime import datetime
try:
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    st.error("Falta la librería 'openpyxl'. Por favor, añádela al archivo requirements.txt")

# Configuración de la página
st.set_page_config(page_title="50 Top Ventas ES", layout="centered")

st.title("🚀 Generador 50 Top Ventas ES")
st.markdown("Carga los archivos en formato **Excel (.xlsx)** para mayor fiabilidad.")

def normalize_sku(series):
    """Limpia y estandariza los SKUs para asegurar el cruce"""
    s = series.astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    return s.apply(lambda x: x.zfill(5) if x.isdigit() else x.upper())

# --- SELECTORES DE ARCHIVOS (Todos XLSX) ---
col1, col2 = st.columns(2)
with col1:
    file_ventas = st.file_uploader("1. VENTAS (Excel)", type=["xlsx"])
    file_stock = st.file_uploader("2. STOCK (Excel)", type=["xlsx"])
    file_excluidos = st.file_uploader("3. EXCLUIDOS (Excel)", type=["xlsx"])
with col2:
    file_tarifa = st.file_uploader("4. TARIFA (Excel)", type=["xlsx"])
    file_feed = st.file_uploader("5. feed_España (Excel)", type=["xlsx"])
    file_ean_extra = st.file_uploader("6. EANs (Excel)", type=["xlsx"])

if st.button("🚀 Ejecutar y Descargar Reporte", use_container_width=True):
    if not all([file_ventas, file_stock, file_excluidos, file_tarifa, file_feed, file_ean_extra]):
        st.warning("⚠️ Por favor, sube los 6 archivos en formato Excel.")
    else:
        try:
            with st.spinner('Procesando datos de Excel...'):
                # 1. CARGA DE DATOS
                df_v = pd.read_excel(file_ventas)
                df_v['ORDEN_V'] = range(len(df_v)) # Preservar orden original
                sku_v_name = df_v.columns[0]
                df_v['SKU_JOIN'] = normalize_sku(df_v[sku_v_name])

                # Stock (Col G es índice 6)
                df_stk = pd.read_excel(file_stock)
                df_stk['SKU_JOIN'] = normalize_sku(df_stk.iloc[:, 0])
                stk_col = df_stk.columns[6]
                df_stk[stk_col] = pd.to_numeric(df_stk[stk_col], errors='coerce').fillna(0)
                df_stk = df_stk.drop_duplicates('SKU_JOIN')

                # Excluidos
                df_exc = pd.read_excel(file_excluidos)
                list_exc = normalize_sku(df_exc.iloc[:, 0]).unique().tolist()

                # Tarifa (SKU: E[4], EAN: F[5], PVPR: G[6])
                df_t = pd.read_excel(file_tarifa)
                df_t_clean = df_t.iloc[:, [4, 5, 6]].copy()
                df_t_clean.columns = ['SKU_JOIN', 'EAN_T', 'PVP_T']
                df_t_clean['SKU_JOIN'] = normalize_sku(df_t_clean['SKU_JOIN'])
                df_t_clean = df_t_clean.drop_duplicates('SKU_JOIN')

                # Feed España (MPN: M[12], Price: Q[16])
                df_f = pd.read_excel(file_feed)
                df_f_clean = df_f.iloc[:, [12, 16]].copy()
                df_f_clean.columns = ['SKU_JOIN', 'PVP_F']
                df_f_clean['SKU_JOIN'] = normalize_sku(df_f_clean['SKU_JOIN'])

                # EANs Extra (A[0], B[1])
                df_e = pd.read_excel(file_ean_extra)
                df_e_clean = df_e.iloc[:, [0, 1]].copy()
                df_e_clean.columns = ['SKU_JOIN', 'EAN_E']
                df_e_clean['SKU_JOIN'] = normalize_sku(df_e_clean['SKU_JOIN'])

                # 2. FILTRADO (Manteniendo el orden de ventas)
                df_proc = pd.merge(df_v, df_stk[['SKU_JOIN', stk_col]], on='SKU_JOIN', how='inner')
                df_proc = df_proc[df_proc[stk_col] > 5] # Stock > 5
                df_proc = df_proc[~df_proc['SKU_JOIN'].isin(list_exc)] # No Excluidos
                df_proc = df_proc[~df_proc['SKU_JOIN'].str.startswith('V')] # No empieza por V
                df_proc = df_proc.sort_values('ORDEN_V')

                # 3. CRUCES Y RESCATE (CASCADA)
                df_res = pd.merge(df_proc, df_t_clean, on='SKU_JOIN', how='left')
                df_res = pd.merge(df_res, df_f_clean, on='SKU_JOIN', how='left')
                df_res = pd.merge(df_res, df_e_clean, on='SKU_JOIN', how='left')

                df_res['EAN_F'] = df_res['EAN_T'].fillna(df_res['EAN_E'])
                df_res['PVP_F'] = df_res['PVP_T'].fillna(df_res['PVP_F'])

                # 4. PREPARACIÓN FINAL
                top_50 = df_res.head(50).copy()
                final_df = pd.DataFrame({
                    "SKU": top_50['SKU_JOIN'],
                    "EAN": pd.to_numeric(top_50['EAN_F'], errors='coerce').fillna(0).astype(int).astype(str).replace("0", ""),
                    "Título del Producto": top_50.iloc[:, 2],
                    "Familia": top_50.iloc[:, 3],
                    "Subfamilia": top_50.iloc[:, 4],
                    "PVPR": pd.to_numeric(top_50['PVP_F'], errors='coerce').fillna(0)
                })

                st.dataframe(final_df.style.format({"PVPR": "{:.2f} €"}), use_container_width=True)

                # 5. GENERACIÓN DE EXCEL CON FORMATO
                today = datetime.now().strftime('%Y%m%d')
                filename = f"{today}_50TopVentasES.xlsx"
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    final_df.to_excel(writer, index=False, sheet_name='Top50')
                    ws = writer.sheets['Top50']
                    
                    # Estilo Encabezado: Azul, Blanco, Negrita
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Formato Moneda y Texto
                    for row in range(2, len(final_df) + 2):
                        ws[f'F{row}'].number_format = '#,##0.00 €'
                        ws[f'B{row}'].number_format = '0'
                    
                    # Auto-ajuste de columnas
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 20

                st.download_button(f"📥 Descargar {filename}", output.getvalue(), file_name=filename, use_container_width=True)

        except Exception as e:
            st.error(f"Se produjo un error al procesar los archivos Excel: {e}")